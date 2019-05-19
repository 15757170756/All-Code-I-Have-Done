# import sys
# sys.stdin = open('input.txt', 'r')


import openpyxl
from openpyxl import Workbook
import calendar
import os
import sys
from openpyxl.styles import Border, Side, \
    Font, numbers, Alignment

border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
alignment = Alignment(horizontal='center', vertical='center')


def get_datelist(month, start_date):
    date_list = []
    days = calendar.monthrange(2019, month)[1]
    for i in range(start_date, days + 1):
        day = i - start_date
        mod = day % 7
        if mod == 4 or mod == 5 or mod == 6:
            continue
        data = str(month) + '.' + str(i)
        date_list.append(data)

    return date_list

def copy_range(start_row, start_col,
               end_row, end_col, sheet):
    range_selected = []
    for i in range(start_row, end_row + 1, 1):
        row_selected = []
        for j in range(start_col, end_col + 1, 1):
            row_selected.append(sheet.cell(row=i, column=j).value)
        range_selected.append(row_selected)

    return range_selected


# Paste range
# Paste data from copyRange into template sheet
def paste_range(start_row, start_col,
                end_row, end_col,
               sheetReceiving, copiedData):
    countRow = 0
    for i in range(start_row, end_row + 1, 1):
        countCol = 0
        for j in range(start_col, end_col + 1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            # sheetReceiving.cell(row=i, column=j).border = border
            # sheetReceiving.cell(row=i, column=j).number_format = numbers.FORMAT_TEXT
            countCol += 1
        countRow += 1

def set_sheet(start_row, start_col,
              end_row, end_col, sheet):
    for i in range(start_row, end_row + 1, 1):
        countCol = 0
        for j in range(start_col, end_col + 1, 1):
            sheet.cell(row=i, column=j).border = border
            sheet.cell(row=i, column=j).number_format = numbers.FORMAT_TEXT
            sheet.cell(row=i, column=j).alignment = alignment

def create_sheet():
    wb_to = Workbook()
    wb_to.create_sheet('all')
    wb_to.create_sheet('汇总')
    for i in range(701, 711, 1):
        sheet = wb_to.create_sheet(str(i))
    sheets = wb_to.sheetnames
    del_sheet = wb_to.get_sheet_by_name(sheets[0])
    wb_to.remove_sheet(del_sheet)

    return wb_to

date_list = get_datelist(7, 1)
stu_number = [14, 23, 8, 12, 12, 15, 13, 9, 27, 16]
sheet_from_name = '6月.xlsx'
sheet_to_name = '7月.xlsx'

wb_from = openpyxl.load_workbook(sheet_from_name)
wb_to = create_sheet()
sheets = wb_to.sheetnames

for i in range(len(stu_number)):
    sheet_from = wb_from.get_sheet_by_name(sheets[i+2])
    # print(sheet_from['A2'].value)
    sheet_to = wb_to.get_sheet_by_name(sheets[i+2])
    select_range = copy_range(1, 1, stu_number[i] + 2, 4, sheet_from)
    # print(select_range)
    paste_range(1, 1, stu_number[i] + 2, 4, sheet_to, select_range)

    for j in range(len(date_list)):
        sheet_to.cell(row=2, column=5 + j).value = date_list[j]

    sheet_to.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + len(date_list))
    set_sheet(1, 1, stu_number[i] + 2, 4 + len(date_list), sheet_to)

wb_to.save(sheet_to_name)

if __name__ == '__main__':
    pass










































# import sys
# sys.stdin = open('input.txt', 'r')


import openpyxl
from openpyxl import Workbook
import calendar
import os
import sys
from openpyxl.styles import Border, Side, \
    Font, numbers, Alignment

border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
alignment = Alignment(horizontal='center', vertical='center')


def get_datelist(month, start_date):
    date_list = []
    days = calendar.monthrange(2019, month)[1]
    for i in range(start_date, days + 1):
        day = i - start_date
        mod = day % 7
        if mod == 4 or mod == 5 or mod == 6:
            continue
        data = str(month) + '.' + str(i)
        date_list.append(data)

    return date_list

def copy_range(start_row, start_col,
               end_row, end_col, sheet):
    range_selected = []
    for i in range(start_row, end_row + 1, 1):
        row_selected = []
        for j in range(start_col, end_col + 1, 1):
            row_selected.append(sheet.cell(row=i, column=j).value)
        range_selected.append(row_selected)

    return range_selected


# Paste range
# Paste data from copyRange into template sheet
def paste_range(start_row, start_col,
                end_row, end_col,
               sheetReceiving, copiedData):
    countRow = 0
    for i in range(start_row, end_row + 1, 1):
        countCol = 0
        for j in range(start_col, end_col + 1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            # sheetReceiving.cell(row=i, column=j).border = border
            # sheetReceiving.cell(row=i, column=j).number_format = numbers.FORMAT_TEXT
            countCol += 1
        countRow += 1

def set_sheet(start_row, start_col,
              end_row, end_col, sheet):
    for i in range(start_row, end_row + 1, 1):
        countCol = 0
        for j in range(start_col, end_col + 1, 1):
            sheet.cell(row=i, column=j).border = border
            sheet.cell(row=i, column=j).number_format = numbers.FORMAT_TEXT
            sheet.cell(row=i, column=j).alignment = alignment

def create_sheet():
    wb_to = Workbook()
    # wb_to.create_sheet('all')
    # wb_to.create_sheet('汇总')
    for i in range(701, 711, 1):
        sheet = wb_to.create_sheet(str(i))
    sheets = wb_to.sheetnames
    del_sheet = wb_to.get_sheet_by_name(sheets[0])
    wb_to.remove_sheet(del_sheet)

    return wb_to

def get_info(sheets):
    all_info = {}
    sheets_name = sheets.sheetnames
    stu_no = dict()
    for i in range(len(sheets_name)):
        sheet_name = sheets_name[i]
        all_stu = []
        sheet_from = sheets.get_sheet_by_name(sheet_name)
        stuNo = 0
        for row in range(4, 50, 1):
            one_stu_info = []
            for col in range(1, 6, 1):
                one_stu_info.append(sheet_from.cell(row=row, column=col).value)
            if one_stu_info[4] == '√':
                all_stu.append(one_stu_info)
            if one_stu_info[0] != None:
                stuNo += 1
        stu_no[sheet_name] = stuNo
        all_info[sheet_name] = all_stu

    return all_info, stu_no

if __name__ == '__main__':
    sheet_from_name = '6月晚托班报名情况.xlsx'
    sheet_to_name = '6月点名册打印.xlsx'

    wb_to = create_sheet()
    stu_infowb = openpyxl.load_workbook(sheet_from_name)
    all_info, stu_no = get_info(stu_infowb)
    # print(stu_no)
    sheets_name = stu_infowb.sheetnames

    date_list = get_datelist(6, 1)
    info_list = ['序号', '姓名', '性别', '联系电话']

    for i in range(len(sheets_name)):
        sheet_from = stu_infowb.get_sheet_by_name(sheets_name[i])
        sheet_to = wb_to.get_sheet_by_name(sheets_name[i])
        allNo = stu_no[sheets_name[i]]
        take_partNo = len(all_info[sheets_name[i]])
        sheet_from.cell(row=2, column=1).value = '（共{}人，参加{}人）'.format(allNo, take_partNo)
        one_sheet_stuInfo = all_info[sheets_name[i]]

        sheet_to.cell(row=1, column=1).value = '{}点名册（{}人）'.format(int(sheets_name[i]), take_partNo)
        for j in range(len(info_list)):
            sheet_to.cell(row=2, column=1+j).value = info_list[j]

        for j in range(take_partNo):
            sheet_to.cell(row=3+j, column=1).value = j + 1
            for k in range(3):
                sheet_to.cell(row=3+j, column=2+k).value = one_sheet_stuInfo[j][k+1]

        for j in range(len(date_list)):
            sheet_to.cell(row=2, column=5 + j).value = date_list[j]

        sheet_to.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + len(date_list))
        set_sheet(1, 1, take_partNo + 2, 4 + len(date_list), sheet_to)

    stu_infowb.save(sheet_from_name)
    wb_to.save(sheet_to_name)